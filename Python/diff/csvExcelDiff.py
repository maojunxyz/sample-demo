# coding=utf-8
import random

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from http.server import ThreadingHTTPServer, SimpleHTTPRequestHandler


import numpy as np
import configparser
import logging
import os
import http.server
import socket
import concurrent.futures

PORT = 8765

Handler = http.server.SimpleHTTPRequestHandler

# 创建一个logger
logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)

# 创建一个handler，用于写入日志文件
file_handler = logging.FileHandler('./output.log', encoding='utf-8')
file_handler.setLevel(logging.DEBUG)

# 定义handler的输出格式
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)

# 给logger添加handler
logging.getLogger().addHandler(file_handler)


logger.debug(
    '========================================================执行开始了===============================================================')

# 初始化ConfigParser对象
config = configparser.ConfigParser()
config.read('settings.ini', encoding='utf-8')

# 获取section下的option
database_settings = config['Database']
skip_column = config['SkipColumn']

# 获取看起来像是键值对的字符串
nested_setting_str = config['ColumnMapping']['mappings']

def get_user_confirmation(message: str) -> bool:
    response = ""
    while response.lower() not in ["y", "n"]:
        response = input(message).strip()  # 获取用户输入并去除前后空白
        if response.lower() == "y":
            return True
        elif response.lower() == "n":
            return False
        print("无效输入，请输入'Y'或'N'。")  # 提示用户输入有效的响应

# 将字符串解析成字典
def parse_nested_value(nested_str):
    if not nested_str:
        return {}
    """将逗号分隔的key:value对解析为字典"""
    pairs = nested_str.split(',')
    result = {}
    for pair in pairs:
        key, value = pair.split(':')
        result[key.strip()] = value.strip()  # 去除前后空白
    return result

# 读取文件
def read_file(file_path, encoding='utf-8'):
    _, file_extension = os.path.splitext(file_path)
    if file_extension.lower() == '.xlsx':
        return pd.read_excel(file_path).replace(' ', np.nan)  # 将空格替换为NaN
    elif file_extension.lower() == '.csv':
        return pd.read_csv(file_path, encoding=encoding, dtype=str).replace(' ', np.nan)  # 将空格替换为NaN，并将所有数据读取为字符串
    else:
        raise ValueError(f"不支持的文件类型: {file_extension}")

def excel_to_column_name(n):
    """
    Convert a number to an Excel column name.
    """
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string

# 获取本机的 IP 地址
def get_local_ip():
    try:
        sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        sock.connect(('10.255.255.255', 1))
        local_ip = sock.getsockname()[0]
    except socket.error:
        local_ip = '127.0.0.1'

    finally:
        sock.close()  # 关闭套接字
    return local_ip



# 自定义请求处理程序类，禁用控制台web日志输出
class QuietHandler(http.server.SimpleHTTPRequestHandler):
    def log_message(self, format, *args):
        pass


#  new begin
# 解析文件对列表
def parse_file_pairs(file_pairs_str):
    pairs = file_pairs_str.split(',')
    file_pairs = []
    for pair in pairs:
        file1, file2 = pair.split(':')
        file_pairs.append((file1.strip(), file2.strip()))  # 去除前后空白
    return file_pairs

# 比较两组文件
def compare_files(file1, file2, column_mapping, skip_columns,output_folder):
    try:
        excel1 = read_file(file1, encoding='utf-8')
        excel2 = read_file(file2, encoding='utf-8')
    except Exception as e:
        logger.error(f"读取文件失败: {e}")
        print(f"读取文件失败: {e}")
        return

    # 将空值替换为相同的标记
    excel1 = excel1.fillna('')
    excel2 = excel2.fillna('')

    # 检查行数是否一致
    if len(excel1) != len(excel2):
        logger.error("错误：{} 和 {} 的行数不一致。".format(file1, file2))
        print("错误：{} 和 {} 的行数不一致。".format(file1, file2))
        return
    # 创建一个用于存储差异的 DataFrame
    diff_cells = []

    # 解析列名映射关系
    column_mapping = parse_nested_value(nested_setting_str)

    columns = config['SkipColumn']

    # 遍历 Excel1 的列
    for col1 in excel1.columns:
        if col1 in skip_column.get('columns').split(","):
            continue

        # 使用映射关系查找 Excel2 中对应的列
        col2 = column_mapping.get(col1, col1)
        if col2 not in excel2.columns:
            logger.error(f"配置错误：列 '{col2}' 未找到在 Excel2 中")
            print(f"配置错误：列 '{col2}' 未找到在 Excel2 中")
            exit(1)

        if col1 in column_mapping.keys():
            index = list(column_mapping.keys()).index(col1)
            for i in excel1[col1].index:
                equal_condition = ((excel1[col1].get(i) == '0') & (excel2[col2].get(i) == '-1')) or (
                        (excel1[col1].get(i) == '0') & (excel2[col2].get(i) == '-2'))
                if equal_condition:
                    excel1.at[i, col1] = '0'
                    excel2.at[i, col2] = '0'

        # 使用上面的函数获取列名（如果col1是列名）
        if isinstance(col1, str):
            col_name = col1
        else:
            col_name = excel_to_column_name(col1 + 1)  # +1是因为Excel列从1开始，而pandas列从0开始

        # 比较两个列中的值
        diff_mask = excel1[col1] != excel2[col2]
        if diff_mask.any():
            for idx in excel1[diff_mask].index:
                diff_cells.append({
                    'row': idx + 2,
                    'column1': col1,
                    'value1': str(excel1.at[idx, col1]),  # 将NaN转换为字符串
                    'column2': col2,
                    'value2': str(excel2.at[idx, col2]),  # 将NaN转换为字符串
                    'coordinate': f"{excel_to_column_name(excel1.columns.get_loc(col1) + 1)}{idx + 2}"
                })
                logger.info(
                    f"第 {idx + 2} 行, 列 '{col1}' 和 '{col2}' 中的值不同：{excel1.at[idx, col1]} != {excel2.at[idx, col2]}")

    # 将差异写入HTML文件
    # 优化后的HTML生成部分，使用Tailwind CSS
    html_content = """
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>文件差异报告</title>
        <link href="../tailwind.min.css" rel="stylesheet">
        <style>
            .highlight:hover {
                background-color: #f0f0f0;
            }
            .table-auto {
                table-layout: fixed;
                width: 100%; /* 表格宽度为100% */

            }
       .table-cell {
                max-width: 300px; /* 设置单元格最大宽度 */
                overflow: hidden;
                text-overflow: ellipsis;
                word-wrap: break-word; /* 在必要时强制内容换行 */
                overflow-wrap: break-word; /* 兼容性更好的写法 */
                transition: max-width 0.3s ease; /* 添加过渡效果 */
            }
            .table-cell:hover {
                max-width: none; /* 鼠标悬停时取消最大宽度限制 */
                overflow: visible;
                white-space: normal;
            }
        </style>
    </head>
    <body class="bg-gray-100 p-4">
        <div class="max-w-full overflow-auto">
            <h2 class="text-2xl font-bold mb-4 text-center">文件差异报告</h2>
            <table class="table-auto w-full border-collapse border border-gray-200">
                <thead>
                    <tr class="bg-gray-200">
                        <th class="px-4 py-2">序号</th>
                        <th class="px-4 py-2">行号</th>
                        <th class="px-4 py-2">文件1列名</th>
                        <th class="px-4 py-2">文件1单元格值</th>
                        <th class="px-4 py-2">文件2列名</th>
                        <th class="px-4 py-2">文件2单元格值</th>
                        <th class="px-4 py-2">Excel坐标</th>
                    </tr>
                </thead>
                <tbody>
    """

    # 循环生成差异
    # 循环生成差异报告表格内容，并增加行号
    for idx, diff in enumerate(diff_cells):
        html_content += f"""
                    <tr class="text-center {'hover:bg-gray-300' if idx % 2 == 0 else 'hover:bg-gray-200'}">
                        <td class="px-4 py-2">{idx + 1}</td>
                        <td class="px-4 py-2">{diff['row']}</td>
                        <td class="px-4 py-2 table-cell overflow-hidden whitespace-nowrap">{diff['column1']}</td>
                        <td class="px-4 py-2 table-cell overflow-hidden whitespace-nowrap">{diff['value1']}</td>
                        <td class="px-4 py-2 table-cell overflow-hidden whitespace-nowrap">{diff['column2']}</td>
                        <td class="px-4 py-2 table-cell overflow-hidden whitespace-nowrap">{diff['value2']}</td>
                        <td class="px-4 py-2">{diff['coordinate']}</td>
                    </tr>
        """

    html_content += """
                </tbody>
            </table>
            </br>
            <h2>备注</h2>
            <p>序号：统计有多少差异数量</p>
            <p>行号：output.excel或文件1中对应的行号</p>
            <p>文件1列名：output.excel或文件1中对应的列名称</p>
            <p>文件1单元格值：output.excel或文件1中，Excel坐标所对应的列名称</p>
            <p>文件2列名：文件2中对应的列名称</p>
            <p>文件2单元格值：文件2中，行号+通过文件2定位的单元格值</p>
            <p>Excel坐标：output.excel或文件1中,由列坐标+行号组成的唯一索引</p>
        </div>
    </body>
    </html>
    """

    html_report_file = os.path.join(output_folder,
                                    f"diff_report_{os.path.basename(file1).split('.')[0]}_{os.path.basename(file2).split('.')[0]}")

    # 将生成的HTML写入文件
    with open(html_report_file+".html", 'w', encoding='utf-8') as html_file:
        html_file.write(html_content)

    # 日志记录部分可以保持不变
    logger.info(f"HTML报告已生成：{html_report_file},已发现差异数量：%d",len(diff_cells))
    print(f"HTML报告已生成：{html_report_file}.html,已发现差异数量：",len(diff_cells))

    # ###################################begin
    #
    # for row in range(len(excel1)):
    #     for col in excel1.columns:
    #         if col in excel2.columns and excel1.at[row, col] != excel2.at[row, col]:
    #             diff_cells.append((row, col))
    #
    # def col_idx_to_name(idx):
    #     name = ''
    #     while idx > 0:
    #         idx, remainder = divmod(idx - 1, 26)
    #         name = chr(65 + remainder) + name
    #     return name
    # # 生成 HTML 内容
    # html_content = """
    # <!DOCTYPE html>
    # <html>
    # <head>
    #     <title>Excel Comparison</title>
    #     <link href="../tailwind.min.css" rel="stylesheet">
    #     <style>
    #         .highlight {{
    #             background-color: #FFFF00;
    #         }}
    #         .search-highlight {{
    #             background-color: #FF0000 !important;
    #         }}
    #         .coord {{
    #             font-size: 0.75em;
    #             color: gray;
    #         }}
    #     </style>
    # </head>
    # <body class="p-4">
    #     <h1 class="text-2xl font-bold mb-4 text-center">Excel差异定位坐标矩阵</h1>
    #     <div class="flex justify-center items-center">
    #         <input type="text" id="coordInput" placeholder="输入Excel坐标 (例如, C4)"
    #             class="border p-2 mb-4" oninput="highlightCell()">
    #     </div>
    #     <div class="overflow-auto">
    #         <table class="min-w-full bg-white border border-gray-300">
    #             <thead>
    #                 {}
    #             </thead>
    #             <tbody>
    #                 {}
    #             </tbody>
    #         </table>
    #     </div>
    #     <script>
    #         function highlightCell() {{
    #             var input = document.getElementById("coordInput").value.toUpperCase();
    #             var cells = document.querySelectorAll("td");
    #             cells.forEach(function(cell) {{
    #                 if (cell.classList.contains("highlight")) {{
    #                     cell.style.backgroundColor = "#FFFF00";
    #                 }} else {{
    #                     cell.style.backgroundColor = "";
    #                 }}
    #                 cell.classList.remove("search-highlight");
    #             }});
    #             if (input) {{
    #                 var targetCell = document.querySelector("[data-coord='" + input + "']");
    #                 if (targetCell) {{
    #                     targetCell.classList.add("search-highlight");
    #                     targetCell.scrollIntoView({{ behavior: 'smooth', block: 'center' }});
    #                 }}
    #             }}
    #         }}
    #     </script>
    # </body>
    # </html>
    # """
    #
    # # 生成表头，添加列坐标
    # header_html = '<tr class="bg-gray-200"><th class="border px-4 py-2 bg-gray-100"></th>'
    # for col_index in range(len(excel1.columns)):
    #     col_letter = excel_to_column_name(col_index + 1)
    #     header_html += f'<th class="border px-4 py-2">{col_letter}</th>'
    # header_html += '</tr>'
    #
    # # 生成表格内容，添加行坐标
    # body_html = ''
    # for row in range(len(excel1)):
    #     row_html = f'<tr><td class="border px-4 py-2 bg-gray-100 font-bold">{row + 1}</td>'
    #     for col_index, col in enumerate(excel1.columns):
    #         cell_value = excel1.at[row, col]
    #         col_letter = excel_to_column_name(col_index + 1)
    #         cell_coord = f'{col_letter}{row + 2}'  # 数据行从2开始
    #         cell_class = 'highlight' if (row, col) in diff_cells else ''
    #         row_html += f'<td data-coord="{cell_coord}" class="border px-4 py-2 {cell_class}">{cell_value} <span class="coord">({cell_coord})</span></td>'
    #     row_html += '</tr>'
    #     body_html += row_html
    #
    # # 合成最终的 HTML 内容
    # html_content = html_content.format(header_html, body_html)
    #
    # # 将 HTML 内容写入文件
    # # html_file = os.path.join(output_folder, name)
    # html_file = html_report_file+"-EXCEL"+".html"
    #
    #
    # with open(html_file, 'w', encoding='utf-8') as file:
    #     file.write(html_content)
    #
    # print(f'HTML report generated: {html_file}')
    # ####################################end

    # 将高亮的 Excel 文件写入到 output.xlsx
    wb = Workbook()
    ws = wb.active

    # 创建一个用于存储高亮信息的 DataFrame
    highlighted_excel = excel1.copy()
    if len(diff_cells) > 0:
        print("比对完成：存在差异正在生成差异Excel文件")
        highlighted_excel.to_excel(html_report_file+'.xlsx', index=False, header=True)

        wb = load_workbook(html_report_file+'.xlsx')
        ws = wb.active

        yellow_fill = PatternFill(start_color='FFFC00', end_color='FFFC00', fill_type='solid')

        for col1 in highlighted_excel.columns:
            if col1 in skip_column.get('columns').split(","):
                continue
            col1_idx = highlighted_excel.columns.get_loc(col1)
            for idx, val in enumerate(highlighted_excel.iloc[:, col1_idx]):
                col2 = column_mapping.get(col1, col1)
                col2_idx = excel2.columns.get_loc(col2)
                if val != excel2.iloc[idx, col2_idx]:
                    ws.cell(row=idx + 2, column=col1_idx + 1).fill = yellow_fill
        wb.save(html_report_file+'.xlsx')
        logger.info(html_report_file+"差异结果生成完成")
        print(html_report_file+"差异结果生成完成")
    else:
        logger.info("比对完成：内容一致跳过生成Excel差异文件")
        print("比对完成：内容一致跳过生成Excel差异文件")


# 创建 output 文件夹
output_folder = 'output'
if not os.path.exists(output_folder):
    os.makedirs(output_folder)

# 主程序开始
file_pairs_str = database_settings.get('files')
file_pairs = parse_file_pairs(file_pairs_str)

# 解析列名映射关系
column_mapping = parse_nested_value(nested_setting_str)

skip_columns = skip_column.get('columns').split(",")

# 使用线程池并行处理文件对
with concurrent.futures.ThreadPoolExecutor() as executor:
    futures = [executor.submit(compare_files, file1, file2, column_mapping, skip_columns,output_folder)
    for file1, file2 in file_pairs]

    # 等待所有任务完成
    concurrent.futures.wait(futures)

# 获取本地IP地址并启动HTTP服务器
local_ip = get_local_ip()
report_url = f"http://{local_ip}:{PORT}/output"

def run(server_class=ThreadingHTTPServer, handler_class=SimpleHTTPRequestHandler):
    server_address = ('', PORT)
    httpd = server_class(server_address, handler_class)
    print(f"结果在线访问地址： {report_url}")
    httpd.serve_forever()

run()

